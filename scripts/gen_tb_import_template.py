# -*- coding: utf-8 -*-
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "/Users/xiayu/Claude-project/leapex/docs/templates/期初试算表导入模板_EasybookX.xlsx"
os.makedirs(os.path.dirname(OUT), exist_ok=True)

FNAME = "Arial"
NAVY = "1F4E8C"; LBLUE = "DCE6F4"; GREY = "F2F2F2"; AMBER = "FBF3EC"; GREEN = "E7F4EA"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '#,##0.00;(#,##0.00);"-"'
RATE = '0.0000'

def hfont(sz=11, b=True, color="FFFFFF"): return Font(name=FNAME, size=sz, bold=b, color=color)
def f(sz=10, b=False, color="000000", it=False): return Font(name=FNAME, size=sz, bold=b, color=color, italic=it)
def fill(c): return PatternFill("solid", fgColor=c)
def center(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def left(wrap=True): return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
def right(): return Alignment(horizontal="right", vertical="center")

wb = Workbook()

# ───────── Sheet 1 · 导入说明 ─────────
s1 = wb.active; s1.title = "导入说明"
s1.sheet_view.showGridLines = False
s1.column_dimensions['A'].width = 4
s1.column_dimensions['B'].width = 16
s1.column_dimensions['C'].width = 92
s1["B1"] = "EasybookX · 期初试算表（Opening TB）导入模板"
s1["B1"].font = Font(name=FNAME, size=16, bold=True, color=NAVY)
s1.merge_cells("B1:C1")
s1["B2"] = "适用：初始化建账 · 香港中小企业 SME · 复式记账 · 币种 HKD（支持外币户备查）"
s1["B2"].font = f(10, color="666666"); s1.merge_cells("B2:C2")

row = 4
def section(title):
    global row
    s1[f"B{row}"] = title; s1[f"B{row}"].font = Font(name=FNAME, size=12, bold=True, color=NAVY)
    s1[f"B{row}"].fill = fill(LBLUE); s1[f"C{row}"].fill = fill(LBLUE)
    s1.merge_cells(f"B{row}:C{row}"); row += 1

def line(k, v):
    global row
    s1[f"B{row}"] = k; s1[f"B{row}"].font = f(10, b=True); s1[f"B{row}"].alignment = Alignment(vertical="top")
    s1[f"C{row}"] = v; s1[f"C{row}"].font = f(10); s1[f"C{row}"].alignment = left()
    row += 1

section("一、填写规则")
rules = [
 ("1", "仅在【初始化建账】时导入一次；之后每月期初余额由系统自动滚转（上月期末 = 本月期初），无需重复导入。"),
 ("2", "只填【永久科目】：资产 A / 负债 L / 权益 E；损益类（收入 I / 费用 X）期初为 0，无需填写。"),
 ("3", "科目代码必须与系统【科目表 COA】完全一致（单一来源）；只填末级可记账科目（postable）。"),
 ("4", "按正常余额方向填：资产 / 费用类填【借方 Dr】；负债 / 权益 / 收入类填【贷方 Cr】。"),
 ("5", "备抵科目（累计折旧 1660 / 16xx、坏账准备 1231）类别仍为 A，但余额填【贷方 Cr】（抵减资产）。"),
 ("6", "外币账户：本位币金额填 借/贷方 列；同时在 原币种 / 原币金额 / 期初汇率 列填写备查（如 USD 户）。"),
 ("7", "【借方合计 = 贷方合计】（试算平衡）方可完成初始化；不平衡将被拦截。"),
 ("8", "留存收益(b/f) 通常为轧平数 = 资产 − 负债 − 股本 − 其他权益（上年累计结果）。"),
]
for k, v in rules: line(k, v)
row += 1

section("二、字段说明")
s1[f"B{row}"]="字段"; s1[f"C{row}"]="说明（★ 必填）"
for c in ("B","C"): s1[f"{c}{row}"].font=f(10,b=True); s1[f"{c}{row}"].fill=fill(GREY)
row+=1
fields = [
 ("科目代码 ★","与 COA 一致的末级科目代码，如 1002-01、1122、3101"),
 ("科目名称","科目中/英文名（与 COA 一致，便于核对）"),
 ("类别 ★","A 资产 / L 负债 / E 权益（下拉选择；损益类不填）"),
 ("借方期初 Dr ★","本位币(HKD)借方期初余额；资产/费用类填此列"),
 ("贷方期初 Cr ★","本位币(HKD)贷方期初余额；负债/权益/收入类与备抵科目填此列"),
 ("原币种","外币户填，如 USD（非外币留空）"),
 ("原币金额","外币户原币金额，如 20000"),
 ("期初汇率","外币户开账日汇率，如 7.8000"),
 ("备注","可选，如 汇丰往来 / 上年累计 / 赊销客户 等"),
]
for k,v in fields:
    s1[f"B{row}"]=k; s1[f"B{row}"].font=f(10,b=True); s1[f"B{row}"].alignment=Alignment(vertical="top")
    s1[f"C{row}"]=v; s1[f"C{row}"].font=f(10); s1[f"C{row}"].alignment=left()
    row+=1
row+=1

section("三、操作步骤")
for v in ["① 在【期初试算表（填写）】页录入贵公司期初余额（可参考【填写示例-兴华贸易】页）。",
          "② 科目代码可查【科目代码参考】页；填完确认底部「借贷平衡」显示 ✓。",
          "③ 保存本文件 → EasybookX 初始化页『导入期初试算表』上传 → 系统校验平衡后完成建账。"]:
    s1[f"B{row}"]=""; s1[f"C{row}"]=v; s1[f"C{row}"].font=f(10); s1[f"C{row}"].alignment=left(); row+=1

# ───────── 通用：填写页构造 ─────────
HEADERS = ["科目代码\nCode","科目名称\nAccount","类别\nA/L/E","借方期初 Dr\n(HKD)","贷方期初 Cr\n(HKD)",
           "原币种\nFX","原币金额\nFX Amt","期初汇率\nRate","备注 Remark"]
WIDTHS = [13, 40, 8, 16, 16, 8, 14, 11, 26]

def build_sheet(ws, data_rows, n_blank=0, title=None, note=None):
    ws.sheet_view.showGridLines = False
    r = 1
    if title:
        ws.cell(r,1,title).font = Font(name=FNAME, size=13, bold=True, color=NAVY); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=9); r+=1
    if note:
        c=ws.cell(r,1,note); c.font=f(9,color="7A3E14",it=True); c.fill=fill(AMBER); c.alignment=left(); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=9); r+=1
    hdr = r
    for j,h in enumerate(HEADERS, start=1):
        c = ws.cell(hdr, j, h); c.font = hfont(10); c.fill = fill(NAVY); c.alignment = center(wrap=True); c.border = border
    ws.row_dimensions[hdr].height = 30
    for j,w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[chr(64+j)].width = w
    first = hdr+1
    rr = first
    for d in data_rows:
        for j in range(1,10):
            c = ws.cell(rr, j); c.border = border; c.font = f(10)
            val = d[j-1] if j-1 < len(d) else None
            c.value = val
            if j in (4,5,7): c.number_format = MONEY; c.alignment = right()
            elif j==8: c.number_format = RATE; c.alignment = right()
            elif j==3: c.alignment = center()
            else: c.alignment = left() if j in (2,9) else center()
        rr += 1
    for _ in range(n_blank):
        for j in range(1,10):
            c = ws.cell(rr, j); c.border = border; c.font=f(10)
            if j in (4,5,7): c.number_format = MONEY; c.alignment=right()
            elif j==8: c.number_format=RATE; c.alignment=right()
            elif j==3: c.alignment=center()
            else: c.alignment=left() if j in (2,9) else center()
        rr += 1
    last = rr-1
    # 合计行
    tot = rr
    ws.cell(tot,1,"合计 Total").font=f(10,b=True); ws.merge_cells(start_row=tot,start_column=1,end_row=tot,end_column=3)
    ws.cell(tot,1).alignment=right(); ws.cell(tot,1).fill=fill(GREY)
    for j in (1,2,3,6,7,8,9): ws.cell(tot,j).fill=fill(GREY); ws.cell(tot,j).border=border
    dD=ws.cell(tot,4,f"=SUM(D{first}:D{last})"); dC=ws.cell(tot,5,f"=SUM(E{first}:E{last})")
    for c in (dD,dC): c.font=f(11,b=True,color=NAVY); c.number_format=MONEY; c.alignment=right(); c.fill=fill(GREY); c.border=border
    ws.cell(tot,6).border=border
    # 平衡校验行
    chk = tot+1
    ws.cell(chk,1,"借贷平衡校验").font=f(10,b=True); ws.merge_cells(start_row=chk,start_column=1,end_row=chk,end_column=3); ws.cell(chk,1).alignment=right()
    formula = f'=IF(ABS(D{tot}-E{tot})<0.01,"✓ 借贷平衡 OK","✗ 不平衡 差额 "&TEXT(D{tot}-E{tot},"#,##0.00"))'
    cc=ws.cell(chk,4,formula); ws.merge_cells(start_row=chk,start_column=4,end_row=chk,end_column=6)
    cc.font=f(11,b=True); cc.alignment=center()
    # 条件着色不便，给绿色底纹提示
    for j in range(4,7): ws.cell(chk,j).fill=fill(GREEN)
    # 类别下拉
    dv = DataValidation(type="list", formula1='"A,L,E"', allow_blank=True); ws.add_data_validation(dv)
    dv.add(f"C{first}:C{last}")
    ws.freeze_panes = ws.cell(first,1)
    return tot, chk

# ───────── Sheet 2 · 填写（空模板）─────────
s2 = wb.create_sheet("期初试算表（填写）")
build_sheet(s2, data_rows=[], n_blank=30,
            title="期初试算表（填写）· 截至建账日（如 2025-04-01）",
            note="★ 请在下方录入贵公司期初余额；类别列下拉选 A/L/E；填完确认底部「借贷平衡 OK」。损益类(I/X)不填。")

# ───────── Sheet 3 · 填写示例（兴华贸易）─────────
EX = [
 ["1001","Cash on hand 库存现金","A",5000,None,None,None,None,""],
 ["1002-01","Bank — HSBC Current 汇丰往来","A",285000,None,None,None,None,"HKD"],
 ["1002-03","Bank — Bank of China 中行","A",156000,None,"USD",20000,7.80,"外币户 USD 20,000×7.80"],
 ["1122","Trade receivables 应收账款","A",180000,None,None,None,None,""],
 ["1231","Allowance for doubtful debts 坏账准备","A",None,8000,None,None,None,"备抵科目(贷方)"],
 ["1123","Prepayments 预付款项","A",24000,None,None,None,None,""],
 ["1605","Motor vehicles 汽车(成本)","A",320000,None,None,None,None,""],
 ["1603","Office equipment 办公设备(成本)","A",60000,None,None,None,None,""],
 ["1604","Computer equipment 电脑设备(成本)","A",45000,None,None,None,None,""],
 ["1615","Accum. dep — motor 累计折旧·汽车","A",None,90000,None,None,None,"备抵科目(贷方)"],
 ["1613","Accum. dep — office 累计折旧·办公","A",None,18000,None,None,None,"备抵科目(贷方)"],
 ["1614","Accum. dep — computer 累计折旧·电脑","A",None,22000,None,None,None,"备抵科目(贷方)"],
 ["2202","Trade payables 贸易应付款","L",None,95000,None,None,None,""],
 ["2242","Accruals 应计费用","L",None,32000,None,None,None,""],
 ["2221","Profits tax payable 应付利得税","L",None,41000,None,None,None,"上年度利得税"],
 ["2251","Amount due to director 应付董事款","L",None,60000,None,None,None,"董事往来"],
 ["3001","Share capital 股本","E",None,100000,None,None,None,"已发行 100,000 股 @ HKD 1"],
 ["3101","Retained earnings (b/f) 期初留存","E",None,609000,None,None,None,"上年累计·轧平数"],
]
s3 = wb.create_sheet("填写示例-兴华贸易")
build_sheet(s3, data_rows=EX, n_blank=0,
            title="填写示例 · 兴华国际贸易（截至 2025-04-01）· 仅供参考",
            note="示例：借贷各 1,075,000.00 平衡。备抵科目(1231/161x)类别 A 但填贷方；外币户(1002-03)填本位币+原币/汇率备查。")

# ───────── Sheet 4 · 科目代码参考 ─────────
s4 = wb.create_sheet("科目代码参考")
s4.sheet_view.showGridLines=False
s4.cell(1,1,"科目代码参考（常用永久科目 · 与系统 COA 一致）").font=Font(name=FNAME,size=12,bold=True,color=NAVY)
s4.merge_cells("A1:E1")
refh=["代码 Code","科目名称 Account","类别","正常余额","说明"]
for j,h in enumerate(refh,1):
    c=s4.cell(2,j,h); c.font=hfont(10); c.fill=fill(NAVY); c.alignment=center(); c.border=border
for j,w in enumerate([13,42,8,10,30],1): s4.column_dimensions[chr(64+j)].width=w
REF=[
 ["1001","Cash on hand 库存现金","A","Dr",""],
 ["1002-01","Bank — HSBC Current 汇丰往来","A","Dr","银行往来(HKD)"],
 ["1002-02","Bank — Standard Chartered 渣打储蓄","A","Dr",""],
 ["1002-03","Bank — Bank of China 中国银行","A","Dr","可作外币户"],
 ["1122","Trade receivables 应收账款","A","Dr",""],
 ["1123","Prepayments 预付款项","A","Dr",""],
 ["1231","Allowance for doubtful debts 坏账准备","A","Cr","备抵(填贷方)"],
 ["1251","Amount due from director 应收董事款","A","Dr",""],
 ["1603","Office equipment 办公设备(成本)","A","Dr",""],
 ["1604","Computer equipment 电脑设备(成本)","A","Dr",""],
 ["1605","Motor vehicles 汽车(成本)","A","Dr",""],
 ["1613","Accum. dep — office 累计折旧·办公","A","Cr","备抵(填贷方)"],
 ["1614","Accum. dep — computer 累计折旧·电脑","A","Cr","备抵(填贷方)"],
 ["1615","Accum. dep — motor 累计折旧·汽车","A","Cr","备抵(填贷方)"],
 ["1660","Accum. dep — PPE 物业厂房设备累折","A","Cr","备抵(填贷方)"],
 ["2202","Trade payables 贸易应付款","L","Cr",""],
 ["2242","Accruals 应计费用","L","Cr",""],
 ["2221","Profits tax payable 应付利得税","L","Cr",""],
 ["2251","Amount due to director 应付董事款","L","Cr","董事往来"],
 ["2502","Bank loans - current 银行贷款(流动)","L","Cr",""],
 ["2801","Bank loans - non-current 银行贷款(非流动)","L","Cr",""],
 ["3001","Share capital 股本","E","Cr",""],
 ["3101","Retained earnings 保留盈余(亏损)","E","Cr","期初留存 b/f"],
 ["3102","Exchange reserve 外币折算储备","E","Cr","列报折算差额"],
]
rr=3
for d in REF:
    for j in range(1,6):
        c=s4.cell(rr,j,d[j-1]); c.border=border; c.font=f(10)
        c.alignment=center() if j in (1,3,4) else left()
    rr+=1
s4.freeze_panes="A3"

from openpyxl.workbook.properties import CalcProperties
wb.calculation = CalcProperties(fullCalcOnLoad=True)
wb.save(OUT)
print("saved:", OUT)
